import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import csv

from exception import FileException, DataException


def read_file(file_path: str):

    reader_by_ext = {
        'xls': read_from_excel,
        'xlsx': read_from_excel,
        'csv': read_from_csv,
    }

    ext = file_path.split('.')[-1]
    if ext not in reader_by_ext:
        raise FileException(f"\"{ext}\" extension not allowed. Only {reader_by_ext.keys()} accepted.")

    headers, data = reader_by_ext.get(ext)(file_path)
    return headers, data


def read_from_excel(file_path):
    body_lists = []
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    # Better use this expression to generate a generator, to save memory
    headers = [value for value in next(sheet.iter_rows(min_row=1, max_row=1, min_col=1, values_only=True))]
    body_generator = sheet.iter_rows(min_row=2, min_col=1, values_only=True)
    for row in body_generator:
        body_lists.append(row)
    return headers, body_lists


def write_to_excel(output_file_path, data_rows, tab_title, new_sheet=False):

    if new_sheet:
        wb = load_workbook(output_file_path)
        ws1 = wb.create_sheet(tab_title)
    else:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = tab_title

    for row in data_rows:
        ws1.append(row)

    wb.save(output_file_path)


def read_from_csv(file_path, encoding='utf-8'):
    body_lists = []
    with open(file_path, 'r', encoding=encoding) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        headers = next(csv_reader)
        for row in csv_reader:
            body_lists.append(row)

    return headers, body_lists


def write_to_csv(output_file_path, data_rows):
    with open(output_file_path, 'w', newline='', encoding='utf-8') as imp_cus_fh:
        imp_cus_csv_writer = csv.writer(imp_cus_fh, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        for data_row in data_rows:
            imp_cus_csv_writer.writerow(data_row)


def upload_file(post_request, file_save_path, file_type):
    make_sure_path_exists(file_save_path)
    file = post_request.files[file_type]
    if not file.filename:
        raise DataException("No file selected. Please try again.")
    file.save(os.path.join(file_save_path, file.filename))
    return os.path.join(file_save_path, file.filename)


def make_sure_path_exists(directory):
    if not os.path.exists(directory):
        os.mkdir(directory)


def acceptable_text_filename(filename: str):
    if filename.endswith('csv') or filename.endswith('xlsx'):
        return True
    else:
        return False


# =================== docx tool =====================
def paragraph_replace_text(paragraph, regex, replace_str):
    """Return `paragraph` after replacing all matches for `regex` with `replace_str`.

    `regex` is a compiled regular expression prepared with `re.compile(pattern)`
    according to the Python library documentation for the `re` module.
    """
    # --- a paragraph may contain more than one match, loop until all are replaced ---
    while True:
        text = paragraph.text
        match = regex.search(text)
        if not match:
            break

        # --- when there's a match, we need to modify run.text for each run that
        # --- contains any part of the match-string.
        runs = iter(paragraph.runs)
        start, end = match.start(), match.end()

        # --- Skip over any leading runs that do not contain the match ---
        for run in runs:
            run_len = len(run.text)
            if start < run_len:
                break
            start, end = start - run_len, end - run_len

        # --- Match starts somewhere in the current run. Replace match-str prefix
        # --- occurring in this run with entire replacement str.
        run_text = run.text
        run_len = len(run_text)
        run.text = "%s%s%s" % (run_text[:start], replace_str, run_text[end:])
        end -= run_len  # --- note this is run-len before replacement ---

        # --- Remove any suffix of match word that occurs in following runs. Note that
        # --- such a suffix will always begin at the first character of the run. Also
        # --- note a suffix can span one or more entire following runs.
        for run in runs:  # --- next and remaining runs, uses same iterator ---
            if end <= 0:
                break
            run_text = run.text
            run_len = len(run_text)
            run.text = run_text[end:]
            end -= run_len

    # --- optionally get rid of any "spanned" runs that are now empty. This
    # --- could potentially delete things like inline pictures, so use your judgement.
    # for run in paragraph.runs:
    #     if run.text == "":
    #         r = run._r
    #         r.getparent().remove(r)

    return paragraph


def doc2pdf(doc_path, dest_path=None):
    if not dest_path:
        dest_path = os.path.dirname(doc_path)

    os_name = os.uname().sysname
    if os_name == 'Darwin':
        import docx2pdf
        docx2pdf.convert(doc_path)
    elif os_name == 'Linux':
        """
            convert a doc/docx document to pdf format (linux only, requires libreoffice)
            :param doc: path to document
            """
        import subprocess
        cmd = f'libreoffice --convert-to pdf --outdir {dest_path}'.split() + [doc_path]
        p = subprocess.Popen(cmd, stderr=subprocess.PIPE, stdout=subprocess.PIPE)
        p.wait(timeout=10)
        stdout, stderr = p.communicate()
        # stderr = stderr.decode('utf-8')
        if stderr:
            stderr_list = stderr.split(b'\n')
            for _stderr in stderr_list:
                if b'warning' in _stderr or not _stderr:
                    continue
                raise subprocess.SubprocessError(stderr)


if __name__ == '__main__':
    read_from_excel('../tmp/test.xlsx')
